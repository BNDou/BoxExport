#!/usr/bin/env python
# coding=utf-8
'''
FilePath     : /BoxExport/BoxExport.py
Description  :  
Author       : BNDou
Date         : 2025-10-28 21:32:19
LastEditTime : 2026-06-22 23:21:24
'''
import os
import sys
import time
import math
import threading
import webbrowser
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from typing import List, Tuple, Optional, Callable, Dict
from datetime import datetime, date

try:
    import xlrd
except ImportError as e:
    print("缺少依赖 xlrd，请先安装：pip install xlrd==1.2.0", file=sys.stderr)
    raise

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
except ImportError as e:
    print("缺少依赖 openpyxl，请先安装：pip install openpyxl", file=sys.stderr)
    raise

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError as e:
    print("缺少依赖 reportlab，请先安装：pip install reportlab", file=sys.stderr)
    raise

# PyInstaller/auto-py-to-exe 打包时，openpyxl 某些内部 writer 模块可能不会被自动收集。
# 下面的显式导入用于作为打包钩子，运行时若不存在也不会报错。
try:
    import openpyxl.cell._writer  # type: ignore
    import openpyxl.workbook._writer  # type: ignore
    import openpyxl.worksheet._writer  # type: ignore
except Exception:
    pass


def ensure_output_dir(script_dir: str) -> str:
    output_dir = os.path.join(script_dir, "output")
    if not os.path.isdir(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    return output_dir


def pick_data_directory(initial_dir: Optional[str] = None) -> Optional[str]:
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askdirectory(title="选择数据源文件夹 data", initialdir=initial_dir or os.getcwd())
    root.update()
    if not path:
        return None
    return path


def list_xls_files(folder: str) -> List[str]:
    files = []
    for name in os.listdir(folder):
        if name.lower().endswith(".xls") and not name.startswith("~$"):
            files.append(os.path.join(folder, name))
    # 按文件名中的数字顺序排序（例如 000013.XLS -> 13）
    def key_func(p: str) -> Tuple[int, str]:
        base = os.path.basename(p)
        stem = os.path.splitext(base)[0]
        try:
            return (int(stem.lstrip('0') or '0'), base)
        except ValueError:
            return (10**9, base)

    files.sort(key=key_func)
    return files


def xcell_value(sheet: "xlrd.sheet.Sheet", row: int, col: int):
    cell = sheet.cell(row, col)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate_as_datetime(cell.value, sheet.book.datemode)
            return dt
        except Exception:
            return cell.value
    return cell.value


def read_records_from_xls(xls_path: str) -> List[dict]:
    """
    源文件格式（按列）：
    A 归档号（不导出）
    C 病案号
    D 出院科室
    E 患者姓名
    F 出院日期
    G 入院日期
    H 成像张数
    I 备注（暂不导出）

    假设第2行是表头，从第3行开始为数据；遇到空的"病案号"即停止。
    """
    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(0)

    records: List[dict] = []

    # 从第3行开始（0-based -> 行索引2）
    start_row = 2
    for r in range(start_row, sheet.nrows):
        case_no = xcell_value(sheet, r, 2)  # C列（index=2）
        if case_no in (None, ""):
            # 若病案号空，认为到达数据尾部
            continue

        department = xcell_value(sheet, r, 3)  # D
        patient_name = xcell_value(sheet, r, 4)  # E
        discharge_time = xcell_value(sheet, r, 5)  # F
        admission_time = xcell_value(sheet, r, 6)  # G
        image_count = xcell_value(sheet, r, 7)  # H

        records.append({
            "case_no": case_no,
            "department": department,
            "patient_name": patient_name,
            "discharge_time": discharge_time,
            "admission_time": admission_time,
            "image_count": image_count,
        })

    return records


def extract_box_number_from_filename(path: str) -> Optional[int]:
    base = os.path.basename(path)
    stem = os.path.splitext(base)[0]
    try:
        return int(stem.lstrip('0') or '0')
    except ValueError:
        return None


def _date_to_yyyymmdd(value):
    """
    将日期值统一转换为 YYYYMMDD 格式字符串（无分隔符）。
    支持 datetime/date 对象与字符串。值为空或转换失败时保持原值。
    """
    if value is None or value == "":
        return value
    try:
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y%m%d")
        if isinstance(value, str):
            # 去除所有非数字字符（去掉 -、/、空格等分隔符），只保留 8 位数字
            digits = "".join(ch for ch in value if ch.isdigit())
            if len(digits) >= 8:
                return digits[:8]
            # 不足 8 位数字，转换失败，保持原值
            return value
    except Exception:
        pass
    return value


def build_and_export_workbook(records: List[dict], box_number: Optional[int], sequence_start: int, out_path: str) -> int:
    wb = Workbook()
    ws = wb.active

    center = Alignment(horizontal="center", vertical="center")
    bold_center = Alignment(horizontal="center", vertical="center")
    vcenter_left = Alignment(horizontal="left", vertical="center")
    vcenter_general = Alignment(horizontal="general", vertical="center")
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    # 表头填充颜色
    header_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

    # A1-H1 合并，标题
    ws.merge_cells("A1:H1")
    ws["A1"] = "案卷目录"
    ws["A1"].alignment = center
    ws["A1"].font = Font(name="Microsoft YaHei UI", size=20, bold=True)
    ws.row_dimensions[1].height = 25

    # E2, F2, G2, H2 固定内容
    ws["E2"] = "全宗号"
    ws["F2"] = "120"
    ws["G2"] = "类目"
    ws["H2"] = "BL"
    for addr in ("E2", "F2", "G2", "H2"):
        ws[addr].alignment = center
        ws[addr].font = Font(name="Microsoft YaHei UI", size=16)
    ws.row_dimensions[2].height = 25

    # G3, H3 箱号
    ws["G3"] = "箱号"
    ws["G3"].alignment = center
    ws["G3"].font = Font(name="Microsoft YaHei UI", size=16)
    ws["H3"] = box_number if box_number is not None else ""
    ws["H3"].alignment = center
    ws["H3"].font = Font(name="Microsoft YaHei UI", size=16)
    ws.row_dimensions[3].height = 25

    # 第4行表头
    headers = [
        ("A4", "顺序号"),
        ("B4", "案卷号（病案号）"),
        ("C4", "出院科室"),
        ("D4", "患者姓名"),
        ("E4", "出院日期"),
        ("F4", "入院日期"),
        ("G4", "成像张数"),
        ("H4", "备注"),
    ]
    for addr, text in headers:
        ws[addr] = text
        ws[addr].alignment = bold_center
        ws[addr].font = Font(name="Microsoft YaHei UI", bold=True)
        ws[addr].fill = header_fill

    # 列宽（按要求）
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 13

    # 数据起始行
    current_row = 5
    seq = sequence_start

    for rec in records:
        # 设置数据行行高
        ws.row_dimensions[current_row].height = 13.5
        # A 顺序号（居中）
        ws.cell(row=current_row, column=1, value=seq).alignment = center
        ws.cell(row=current_row, column=1, value=seq).font = Font(name="Microsoft YaHei UI")

        # B-H 数据
        ws.cell(row=current_row, column=2, value=rec.get("case_no"))
        ws.cell(row=current_row, column=3, value=rec.get("department"))
        ws.cell(row=current_row, column=4, value=rec.get("patient_name"))
        # 设置字体和垂直居中
        for col in range(2, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(name="Microsoft YaHei UI")
            cell.alignment = vcenter_left

        # E 出院日期, F 入院日期（统一转为 YYYYMMDD 格式字符串）
        e_val = _date_to_yyyymmdd(rec.get("discharge_time"))
        f_val = _date_to_yyyymmdd(rec.get("admission_time"))

        e_cell = ws.cell(row=current_row, column=5, value=e_val)
        f_cell = ws.cell(row=current_row, column=6, value=f_val)
        # 设置垂直居中对齐
        e_cell.alignment = vcenter_left
        f_cell.alignment = vcenter_left

        # G 成像张数（居中）
        g_cell = ws.cell(row=current_row, column=7, value=rec.get("image_count"))
        g_cell.alignment = center
        g_cell.font = Font(name="Microsoft YaHei UI")

        # H 备注 先留空
        ws.cell(row=current_row, column=8, value="")
        ws.cell(row=current_row, column=8).font = Font(name="Microsoft YaHei UI")

        current_row += 1
        seq += 1

    # 添加表格边框：A1 到 H{最后一行}
    last_row = ws.max_row
    for r in range(1, last_row + 1):
        for c in range(1, 8 + 1):
            ws.cell(row=r, column=c).border = thin_border

    wb.save(out_path)
    # 返回下一个可用顺序号
    return seq


def build_and_export_pdf(records: List[dict], box_number: Optional[int], sequence_start: int, out_path: str) -> int:
    """
    使用reportlab生成PDF格式的案卷目录，A4纸张大小，完全匹配Excel格式
    """
    # 初始化字体变量
    registered_font_name = 'Helvetica'
    registered_bold_font = 'Helvetica-Bold'
    
    # 设置较小的上下边距，确保内容更多
    doc = SimpleDocTemplate(out_path, pagesize=A4, topMargin=6*mm, bottomMargin=10*mm)
    
    # 注册中文字体（优先使用项目fonts目录，再使用系统字体）
    try:
        # 获取项目根目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        fonts_dir = os.path.join(script_dir, "fonts")
        
        # 优先使用项目fonts目录中的MiSans字体
        normal_font = os.path.join(fonts_dir, "MiSans-Normal.ttf")
        bold_font = os.path.join(fonts_dir, "MiSans-Demibold.ttf")
        
        if os.path.exists(normal_font) and os.path.exists(bold_font):
            pdfmetrics.registerFont(TTFont('MiSans-Normal', normal_font))
            pdfmetrics.registerFont(TTFont('MiSans-Demibold', bold_font))
            registered_font_name = 'MiSans-Normal'
            registered_bold_font = 'MiSans-Demibold'
        else:
            # 如果项目目录中没有，则尝试系统字体
            # 尝试微软雅黑
            yahei_normal = r"C:\Windows\Fonts\msyh.ttf"
            yahei_bold = r"C:\Windows\Fonts\msyhbd.ttf"
            
            if os.path.exists(yahei_normal) and os.path.exists(yahei_bold):
                pdfmetrics.registerFont(TTFont('MicrosoftYaHei', yahei_normal))
                pdfmetrics.registerFont(TTFont('MicrosoftYaHei-Bold', yahei_bold))
                registered_font_name = 'MicrosoftYaHei'
                registered_bold_font = 'MicrosoftYaHei-Bold'
            else:
                # 备选方案：微软雅黑复合字体
                yahei_ttc = r"C:\Windows\Fonts\msyh.ttc"
                if os.path.exists(yahei_ttc):
                    pdfmetrics.registerFont(TTFont('MicrosoftYaHei', yahei_ttc))
                    pdfmetrics.registerFont(TTFont('MicrosoftYaHei-Bold', yahei_ttc))
                    registered_font_name = 'MicrosoftYaHei'
                    registered_bold_font = 'MicrosoftYaHei-Bold'
                else:
                    # 最后备选：黑体
                    simhei_font = r"C:\Windows\Fonts\simhei.ttf"
                    if os.path.exists(simhei_font):
                        pdfmetrics.registerFont(TTFont('SimHei', simhei_font))
                        pdfmetrics.registerFont(TTFont('SimHei-Bold', simhei_font))
                        registered_font_name = 'SimHei'
                        registered_bold_font = 'SimHei-Bold'
                    else:
                        # 如果都没有，使用默认字体
                        registered_font_name = 'Helvetica'
                        registered_bold_font = 'Helvetica-Bold'
                        
    except Exception as e:
        # 如果字体注册失败，使用默认字体
        registered_font_name = 'Helvetica'
        registered_bold_font = 'Helvetica-Bold'

    elements = []
    
    # 使用已注册的字体名称
    font_name = registered_font_name
    bold_font = registered_bold_font
    
    # 构建完整的表格数据，包含标题行和所有信息
    table_data = []
    
    # 第一行：合并的标题 "案卷目录"
    table_data.append(['案卷目录', '', '', '', '', '', '', ''])
    
    # 第二行：全宗号信息（空A-D列，E列"全宗号", F列"120", G列"类目", H列"BL"）
    table_data.append(['', '', '', '', '全宗号', '120', '类目', 'BL'])
    
    # 第三行：箱号信息（空A-F列，G列"箱号", H列箱号数字）
    table_data.append(['', '', '', '', '', '', '箱号', box_number if box_number is not None else ''])
    
    # 第四行：表头
    table_data.append(['顺序号', '案卷号 (病案号)', '出院科室', '患者姓名', '出院日期', '入院日期', '成像张数', '备注'])
    
    # 添加数据行
    seq = sequence_start
    for rec in records:
        # 日期统一转为 YYYYMMDD 格式字符串
        discharge_time = _date_to_yyyymmdd(rec.get("discharge_time", ""))
        admission_time = _date_to_yyyymmdd(rec.get("admission_time", ""))
        
        table_data.append([
            str(seq),
            str(rec.get("case_no", "")),
            str(rec.get("department", "")),
            str(rec.get("patient_name", "")),
            str(discharge_time),
            str(admission_time),
            str(rec.get("image_count", "")),
            ""  # 备注留空
        ])
        seq += 1
    
    # 设置所有列的宽度（按Excel比例调整，更紧凑）
    col_widths = [15*mm, 30*mm, 30*mm, 18*mm, 21*mm, 21*mm, 18*mm, 20*mm]
    
    # 创建完整表格
    table = Table(table_data, colWidths=col_widths, repeatRows=4)  # 设置表头重复
    
    # 应用详细样式
    style = TableStyle([
        # 字体设置
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # 整体字体大小
        
        # 第一行标题样式
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTSIZE', (0, 0), (-1, 0), 20),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('SPAN', (0, 0), (7, 0)),  # 合并单元格
        
        # 第二行信息样式
        ('FONTSIZE', (4, 1), (7, 1), 14),
        ('ALIGN', (4, 1), (7, 1), 'CENTER'),
        ('VALIGN', (4, 1), (7, 1), 'MIDDLE'),
        
        # 第三行箱号样式 
        ('FONTSIZE', (6, 2), (7, 2), 14),
        ('ALIGN', (6, 2), (7, 2), 'CENTER'),
        ('VALIGN', (6, 2), (7, 2), 'MIDDLE'),
        
        # 第四行表头样式
        ('FONTNAME', (0, 3), (-1, 3), bold_font),
        ('FONTSIZE', (0, 3), (-1, 3), 11),
        ('ALIGN', (0, 3), (-1, 3), 'CENTER'),
        ('VALIGN', (0, 3), (-1, 3), 'MIDDLE'),
        ('BACKGROUND', (0, 3), (-1, 3), colors.Color(239/255, 239/255, 239/255)),  # #EFEFEF
        
        # 数据行样式
        ('FONTSIZE', (0, 4), (-1, -1), 9),
        ('ALIGN', (0, 4), (0, -1), 'CENTER'),  # 顺序号居中
        ('ALIGN', (6, 4), (6, -1), 'CENTER'),  # 成像张数居中
        ('VALIGN', (0, 4), (-1, -1), 'MIDDLE'),
        
        # 优化边框设置 - 减少边框数量和粗细以减小文件体积
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # 外边框略粗
        
        # 简化行高调整
        ('TOPPADDING', (0, 0), (-1, 0), 0),    # 标题行上方内边距
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12), # 标题行下方内边距
        ('TOPPADDING', (0, 1), (-1, 2), -1),      # 信息行内边距
        ('BOTTOMPADDING', (0, 1), (-1, 2), 5),  # 信息行内边距
        ('TOPPADDING', (0, 3), (-1, 3), 1),      # 表头内边距
        ('BOTTOMPADDING', (0, 3), (-1, 3), 2),  # 表头内边距
        ('TOPPADDING', (0, 4), (-1, -1), 2),     # 数据行上边距减小
        ('BOTTOMPADDING', (0, 4), (-1, -1), 0), # 数据行下边距减小
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # 生成PDF
    doc.build(elements)
    
    return seq


# 科室全称 -> 简写 映射表
DEPARTMENT_ABBR_MAP: Dict[str, str] = {
    "ICU二病区": "ICU二",
    "中医科病区": "中医",
    "产科病区": "产",
    "儿内科病区": "儿",
    "全科医疗科病区": "全科",
    "内分泌科病区": "内分泌",
    "口腔科病区": "口",
    "呼吸内科病区": "呼",
    "妇科病区": "妇",
    "康复科病区": "康",
    "心胸外科病区": "胸外",
    "心血管内科病区": "心内",
    "急诊医学科病区": "急诊医学",
    "新生儿科病区": "新",
    "普通外科一病区": "普一",
    "普通外科三病区": "普三",
    "普通外科二病区": "普二",
    "泌尿外科病区": "泌外",
    "消化内科病区": "消",
    "皮肤科病区": "皮",
    "眼科病区": "眼",
    "神经内科病区": "神内",
    "神经外科病区": "神外",
    "老年病科一病区": "干一",
    "老年病科二病区": "干二",
    "耳鼻喉科病区": "耳",
    "肾病学病区": "肾",
    "肿瘤科病区": "肿",
    "血液内科病区": "血液",
    "血管外科病区": "血外",
    "重症医学科": "重",
    "骨科一病区": "骨一",
    "骨科三病区": "骨三",
    "骨科二病区": "骨二",
}


def _abbreviate_department(dept_name: str) -> str:
    """
    将科室全称转换为简写。
    若科室名不在映射表中，则返回去除首尾空格的原值（兜底处理，保持健壮性）。
    """
    if dept_name is None:
        return ""
    cleaned = str(dept_name).strip()
    return DEPARTMENT_ABBR_MAP.get(cleaned, cleaned)


def merge_duplicate_records(records: List[dict]) -> List[dict]:
    """
    对同一文件内相同案卷号（case_no）的多条记录进行合并。
    合并规则：
    - 入院日期（admission_time）：取所有记录中的最小值（最早日期）
    - 出院日期（discharge_time）：取所有记录中的最大值（最晚日期）
    - 成像张数（image_count）：取所有记录的总和
    - 出院科室（department）：所有记录科室相同则输出该科室全称；
      不同则输出各科室简写并用中文顿号「、」连接
    - 其他字段（patient_name）：取第一条记录的值
    保持合并后记录按首次出现的顺序排列。
    """
    if not records:
        return records

    def _to_date_obj(value):
        """将日期值转换为可比较的 date 对象，失败返回 None。"""
        if value is None or value == "":
            return None
        try:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                digits = "".join(ch for ch in value if ch.isdigit())
                if len(digits) >= 8:
                    return datetime.strptime(digits[:8], "%Y%m%d").date()
        except Exception:
            pass
        return None

    def _to_float(value):
        """将成像张数转为 float，失败返回 0.0。"""
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def _date_obj_to_orig(date_obj, original_value):
        """将 date 对象转回原始值的格式。"""
        if isinstance(original_value, (datetime, date)):
            return date_obj
        return date_obj.strftime("%Y%m%d")

    # 按 case_no 分组，保持首次出现顺序
    groups: Dict[object, List[dict]] = {}
    order: List[object] = []
    for rec in records:
        key = rec.get("case_no")
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(rec)

    merged: List[dict] = []
    for key in order:
        group = groups[key]
        if len(group) == 1:
            merged.append(dict(group[0]))
            continue

        first = dict(group[0])

        # 入院日期取最小值（最早日期）
        admission_dates = [_to_date_obj(r.get("admission_time")) for r in group]
        valid_admission = [d for d in admission_dates if d is not None]
        if valid_admission:
            min_admission = min(valid_admission)
            first["admission_time"] = _date_obj_to_orig(min_admission, group[0].get("admission_time"))

        # 出院日期取最大值（最晚日期）
        discharge_dates = [_to_date_obj(r.get("discharge_time")) for r in group]
        valid_discharge = [d for d in discharge_dates if d is not None]
        if valid_discharge:
            max_discharge = max(valid_discharge)
            first["discharge_time"] = _date_obj_to_orig(max_discharge, group[0].get("discharge_time"))

        # 成像张数取总和（先转 float 求和，再转回 int 如果和是整数）
        total_count = sum(_to_float(r.get("image_count")) for r in group)
        if total_count == int(total_count):
            first["image_count"] = int(total_count)
        else:
            first["image_count"] = total_count

        # 出院科室合并处理：相同则输出全称，不同则输出简写并用中文顿号连接
        departments = []
        for r in group:
            dept = r.get("department")
            if dept is None:
                continue
            dept = str(dept).strip()
            if dept and dept not in departments:
                departments.append(dept)
        if len(departments) == 1:
            first["department"] = departments[0]
        elif len(departments) > 1:
            first["department"] = "、".join(
                _abbreviate_department(d) for d in departments
            )

        merged.append(first)

    return merged


def process_all(
    data_dir: str,
    sequence_start: int,
    output_dir: str,
    export_type: str = "pdf",  # "excel" 或 "pdf"
    on_progress: Optional[Callable[[int, int, str, Dict[str, float]], None]] = None,
    on_log: Optional[Callable[[str], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
) -> None:
    # 确保用户自定义的导出目录存在
    os.makedirs(output_dir, exist_ok=True)

    xls_files = list_xls_files(data_dir)
    if not xls_files:
        raise RuntimeError("所选目录未找到 .XLS 文件。")

    current_seq = sequence_start
    total = len(xls_files)
    start_time = time.time()

    def ftime(sec: float) -> str:
        sec = int(sec)
        h = sec // 3600
        m = (sec % 3600) // 60
        s = sec % 60
        if h > 0:
            return f"{h:02d}:{m:02d}:{s:02d}"
        else:
            return f"{m:02d}:{s:02d}"

    for idx, xls in enumerate(xls_files, start=1):
        file_start = time.time()
        name = os.path.basename(xls)
        if on_log:
            on_log(f"读取: {name}")
        records = read_records_from_xls(xls)
        records = merge_duplicate_records(records)

        box_no = extract_box_number_from_filename(xls)
        
        if export_type == "pdf":
            out_name = f"{os.path.splitext(name)[0]}.pdf"
            out_path = os.path.join(output_dir, out_name)
        else:
            out_name = f"{os.path.splitext(name)[0]}.xlsx"
            out_path = os.path.join(output_dir, out_name)

        if on_log:
            on_log(f"导出: {out_name}")
        
        if export_type == "pdf":
            next_seq = build_and_export_pdf(records, box_no, current_seq, out_path)
        else:
            next_seq = build_and_export_workbook(records, box_no, current_seq, out_path)
        current_seq = next_seq

        file_elapsed = time.time() - file_start
        elapsed = time.time() - start_time
        avg_per_file = elapsed / idx
        remain_files = total - idx
        eta_seconds = max(0.0, avg_per_file * remain_files)
        if on_progress:
            on_progress(
                idx,
                total,
                name,
                {
                    "file_elapsed": file_elapsed,
                    "elapsed": elapsed,
                    "avg_per_file": avg_per_file,
                    "eta_seconds": eta_seconds,
                },
            )

        # 若用户请求停止，则在完成当前文件后终止后续处理
        if should_stop and should_stop():
            if on_log:
                on_log("收到停止指令，已在当前文件完成后终止后续处理。")
            break

    if on_log:
        on_log(f"完成，输出目录: {output_dir}")


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("病案案卷目录批量生成")
        self.root.geometry("820x520+0+0")

        self.data_dir_var = tk.StringVar()
        self.seq_var = tk.IntVar(value=1)
        self.output_dir_var = tk.StringVar(value=os.path.join(os.path.dirname(os.path.abspath(__file__)), "output"))
        self.export_type_var = tk.StringVar(value="pdf")  # 默认导出类型为PDF
        self.status_var = tk.StringVar(value="就绪")
        self.percent_var = tk.StringVar(value="0.00%")
        self.eta_var = tk.StringVar(value="ETA 00:00")
        self.elapsed_var = tk.StringVar(value="已用 00:00")
        self.avg_var = tk.StringVar(value="平均 00:00")

        self._build_ui()

        self._total = 0
        self._idx = 0
        self._start_ts = 0.0
        self._running = False

    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}

        frm_top = ttk.Frame(self.root)
        frm_top.pack(fill=tk.X, **pad)

        ttk.Label(frm_top, text="数据目录:").pack(side=tk.LEFT)
        ent_dir = ttk.Entry(frm_top, textvariable=self.data_dir_var, width=70)
        ent_dir.pack(side=tk.LEFT, padx=6)
        ttk.Button(frm_top, text="浏览…", command=self._browse).pack(side=tk.LEFT)

        frm_seq = ttk.Frame(self.root)
        frm_seq.pack(fill=tk.X, **pad)
        ttk.Label(frm_seq, text="起始顺序号:").pack(side=tk.LEFT)
        spn = ttk.Spinbox(frm_seq, from_=1, to=999999, textvariable=self.seq_var, width=8)
        spn.pack(side=tk.LEFT, padx=6)
        
        # 导出类型选择
        ttk.Label(frm_seq, text="导出格式:").pack(side=tk.LEFT, padx=(12, 0))
        excel_radio = ttk.Radiobutton(frm_seq, text="Excel", variable=self.export_type_var, value="excel")
        excel_radio.pack(side=tk.LEFT)
        pdf_radio = ttk.Radiobutton(frm_seq, text="PDF", variable=self.export_type_var, value="pdf")
        pdf_radio.pack(side=tk.LEFT, padx=(0, 6))
        
        self.btn_start = ttk.Button(frm_seq, text="开始", command=self._start)
        self.btn_start.pack(side=tk.LEFT)
        self.btn_stop = ttk.Button(frm_seq, text="停止", command=self._stop, state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT, padx=6)

        frm_out = ttk.Frame(self.root)
        frm_out.pack(fill=tk.X, **pad)
        ttk.Label(frm_out, text="导出目录:").pack(side=tk.LEFT)
        ent_out = ttk.Entry(frm_out, textvariable=self.output_dir_var, width=70)
        ent_out.pack(side=tk.LEFT, padx=6)
        ttk.Button(frm_out, text="浏览…", command=self._browse_out).pack(side=tk.LEFT)
        ttk.Button(frm_out, text="打开导出目录", command=self._open_out_dir).pack(side=tk.LEFT, padx=6)

        frm_prog = ttk.Frame(self.root)
        frm_prog.pack(fill=tk.X, **pad)
        self.prog = ttk.Progressbar(frm_prog, orient=tk.HORIZONTAL, length=600, mode='determinate', maximum=100)
        self.prog.pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Label(frm_prog, textvariable=self.percent_var, width=8).pack(side=tk.LEFT, padx=6)

        frm_stats = ttk.Frame(self.root)
        frm_stats.pack(fill=tk.X, **pad)
        ttk.Label(frm_stats, textvariable=self.status_var).pack(side=tk.LEFT)
        ttk.Label(frm_stats, textvariable=self.elapsed_var).pack(side=tk.LEFT, padx=12)
        ttk.Label(frm_stats, textvariable=self.avg_var).pack(side=tk.LEFT, padx=12)
        ttk.Label(frm_stats, textvariable=self.eta_var).pack(side=tk.LEFT, padx=12)

        frm_log = ttk.LabelFrame(self.root, text="日志")
        frm_log.pack(fill=tk.BOTH, expand=True, **pad)
        self.txt = tk.Text(frm_log, height=16)
        self.txt.pack(fill=tk.BOTH, expand=True)

        # 右下角版本与作者信息，可点击打开链接
        frm_bottom = ttk.Frame(self.root)
        frm_bottom.pack(fill=tk.X, side=tk.BOTTOM)
        link = tk.Label(
            frm_bottom,
            text="v2026.06.22 by BNDou",
            fg="blue",
            font=("SimHei", 10, "bold"),
            cursor="hand2",
            anchor=tk.E,
        )
        link.pack(side=tk.RIGHT, padx=10, pady=6)
        link.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/BNDou/BoxExport"))

    def _browse(self):
        d = filedialog.askdirectory(title="选择数据源文件夹 data", initialdir=self.data_dir_var.get() or os.getcwd())
        if d:
            self.data_dir_var.set(d)

    def _browse_out(self):
        d = filedialog.askdirectory(title="选择导出目录", initialdir=self.output_dir_var.get() or os.getcwd())
        if d:
            self.output_dir_var.set(d)

    def _open_out_dir(self):
        p = self.output_dir_var.get().strip()
        if not p:
            messagebox.showwarning("提示", "请先选择导出目录。")
            return
        if not os.path.isdir(p):
            messagebox.showwarning("提示", "导出目录不存在，请先生成或选择有效目录。")
            return
        try:
            if sys.platform.startswith('win'):
                os.startfile(p)  # type: ignore
            elif sys.platform == 'darwin':
                subprocess.call(['open', p])
            else:
                subprocess.call(['xdg-open', p])
        except Exception as e:
            messagebox.showerror("错误", f"无法打开目录：{e}")

    def _start(self):
        if self._running:
            return
        data_dir = self.data_dir_var.get().strip()
        if not data_dir or not os.path.isdir(data_dir):
            messagebox.showwarning("提示", "请先选择有效的数据目录。")
            return
        output_dir = self.output_dir_var.get().strip()
        if not output_dir:
            messagebox.showwarning("提示", "请先选择导出目录。")
            return
        seq = int(self.seq_var.get() or 1)
        self._running = True
        self.btn_start.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self.prog['value'] = 0
        self.percent_var.set("0.00%")
        self.txt.delete(1.0, tk.END)
        self.status_var.set("开始处理…")
        self._start_ts = time.time()
        self._stop_requested = False

        def run():
            try:
                process_all(
                    data_dir,
                    seq,
                    output_dir,
                    export_type=self.export_type_var.get(),
                    on_progress=self._on_progress,
                    on_log=self._on_log,
                    should_stop=lambda: self._stop_requested,
                )
                self.root.after(0, lambda: messagebox.showinfo("完成", "处理完成。请在导出目录查看导出文件。"))
            except Exception as e:
                self.root.after(0, lambda e=e: messagebox.showerror("错误", str(e)))
            finally:
                def finish():
                    self._running = False
                    self.btn_start.config(state=tk.NORMAL)
                    self.btn_stop.config(state=tk.DISABLED)
                    self.status_var.set("就绪")
                    self._stop_requested = False
                self.root.after(0, finish)

        threading.Thread(target=run, daemon=True).start()

    def _on_log(self, msg: str):
        def do():
            self.txt.insert(tk.END, msg + "\n")
            self.txt.see(tk.END)
        self.root.after(0, do)

    def _on_progress(self, idx: int, total: int, name: str, stats: Dict[str, float]):
        percent = idx * 100.0 / total if total else 0.0
        def ftime(sec: float) -> str:
            sec = int(sec)
            h = sec // 3600
            m = (sec % 3600) // 60
            s = sec % 60
            if h > 0:
                return f"{h:02d}:{m:02d}:{s:02d}"
            else:
                return f"{m:02d}:{s:02d}"
        def do():
            self.prog['value'] = percent
            self.percent_var.set(f"{percent:0.2f}%")
            self.status_var.set(f"{idx}/{total}  当前: {name}")
            self.elapsed_var.set(f"已用 {ftime(stats.get('elapsed', 0.0))}")
            self.avg_var.set(f"平均 {ftime(stats.get('avg_per_file', 0.0))}")
            self.eta_var.set(f"ETA {ftime(stats.get('eta_seconds', 0.0))}")
        self.root.after(0, do)

    def _stop(self):
        if not self._running:
            return
        self._stop_requested = True
        self.status_var.set("已请求停止：将完成当前文件后终止…")


def main_gui():
    root = tk.Tk()
    App(root)
    root.mainloop()


def _wait_exit():
    try:
        print("输入 0 然后回车以退出程序...")
        while True:
            user_in = input().strip()
            if user_in == "0":
                break
    except Exception:
        pass


if __name__ == "__main__":
    # 默认启动 GUI
    main_gui()
